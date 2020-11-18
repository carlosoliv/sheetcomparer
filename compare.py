from datetime import datetime
import openpyxl
import os
import sys
import warnings

if len(sys.argv) == 1:
    rota = os.getcwd()+"/Deployment 2020 Rota.xlsx"
    avail = os.getcwd()+"/Availability_email_template_2020.xlsx"
elif len(sys.argv) == 3:
    rota = os.getcwd()+'/'+sys.argv[1]
    avail = os.getcwd()+'/'+sys.argv[2]
else:
    print (f'Usage:\npython {sys.argv[0]} <rota.xlsx> <availability.xlsx>')
    sys.exit(1)

try:
    wb_rota = openpyxl.load_workbook(rota, read_only = True, data_only = True)
except:
    print (f'Failure reading rota file \'{rota}\' !')
    sys.exit(1)
sheet_rota = wb_rota.active

try:
    warnings.simplefilter("ignore")
    wb_avail = openpyxl.load_workbook(filename=avail, data_only='True')
    warnings.simplefilter("default")
except:
    print (f'Failure reading availability file \'{avail}\' !')
    sys.exit(1)
sheet_avail = wb_avail.active

# Testing if we have the right files
if sheet_rota['A1'].value != 'POD':
    print (f'The file \'{rota}\' was not recognized as a valid Rota file!')
    sys.exit(1)

if sheet_avail['J2'].value != 'MORNING EMAIL':
    print (f'The file \'{avail}\' was not recognized as a valid Availability Email Template!')
    sys.exit(1)

dict_rota = {}
dict_avail = {}
dict_left = {}

# Getting collumn for today
today = datetime.today().strftime('%Y-%m-%d')

for rows in sheet_rota.iter_rows(min_row=2, max_row=2, min_col=10, max_col=400):
    for cell in rows:
        if today == str(cell.value).split(' ')[0]:
            today_col = str(cell)[-4:-2]

# Getting all engineers in the rota, adding to the dict_rota{}
for cell in sheet_rota.iter_rows(min_row=3, max_row=80, min_col=2, max_col=2):
    if cell[0].value == None or cell[0].value.split()[0][0:4] == 'http':
        break

    # Finding people who left the team
    color_in_hex = cell[0].fill.start_color.index
    if color_in_hex == 'FFFF0000':
        dict_left[cell[0].value.split()[0]] = ''

    # Finding engineer status for today
    eng_col = str(cell[0]).split()[1].split('.')[1][1:-1]
    eng_status = sheet_rota[today_col+eng_col].value
    #print (f'eng: {cell[0].value.split()[0]} | col: {eng_col} | eng_status: {eng_status} | pos-status: {today_col+eng_col}')

    dict_rota[cell[0].value.split()[0]] = eng_status

#for k,v in dict_rota.items():
#    print(f'eng: {k} | status: {v}')

# Getting all engineers in the availability template, adding to dict_avail[]
for cell in sheet_avail.iter_rows(min_row=30, max_row=150, min_col=11, max_col=11):
    if cell[0].value == None:
        break
    
    # Updating engineer status in availability template
    col = str(cell[0]).split()[1][-3:-1]
    if cell[0].value.split()[0] in dict_rota and dict_rota[cell[0].value.split()[0]] == 'X':
        if cell[0].value.split()[0] == 'davidlyn' or cell[0].value.split()[0] == 'jdiv' or cell[0].value.split()[0] == 'niamhf' or cell[0].value.split()[0] == 'bddrysda':
            sheet_avail['M'+col].value = 'In'
        else:
            sheet_avail['M'+col].value = 'Available (WFH)'
    else:
        sheet_avail['M'+col].value = 'Off'

    dict_avail[cell[0].value.split()[0]] = ''

wb_avail.save(avail[:-5]+'-UPDATED.xlsx')

print (f'\nNumber of engineers in the Rota: {len(dict_rota)}\nNumber of engineers in Availability template: {len(dict_avail)}')

avail_today = 0
for k,v in dict_rota.items():
    if v == 'X' and k != 'davidlyn' and k != 'jdiv' and k != 'niamhf' and k != 'bddrysda':
        avail_today += 1
print(f'Engineers available today: {avail_today}')

print(f'\nIn the Availability template but missing in rota:')
for i in dict_avail:
    if i not in dict_rota and i not in dict_left:
        print (f'{i}')

print(f'\nIn the rota but missing in the availability template:')
for i in dict_rota:
    if i not in dict_avail and i not in dict_left:
        print (f'{i}')

if len(dict_left) != 0:
    print (f'\nWarning: these people have left the team!\nPlease remove them from the Availability Template!')
    for i in dict_left:
        print (i)
